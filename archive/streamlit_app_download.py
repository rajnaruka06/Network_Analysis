import os
os.add_dll_directory(r"C:\Program Files\GTK3-Runtime Win64\bin")

import streamlit as st
from weasyprint import HTML
import pandas as pd
import networkx as nx
from pyvis.network import Network # type: ignore
from networkx.algorithms import community
import matplotlib.colors as mcolors
import numpy as np
from openpyxl import load_workbook
import base64


def _read_csv(upload_file):
    df = pd.read_csv(upload_file)
    df.columns = ['source', 'target']
    df = df.astype(str)
    for col in df.columns:
        df[col] = df[col].str.strip()
    df.dropna(inplace=True)
    df.drop_duplicates(inplace=True)
    return df

def _read_excel(uploaded_file):
    wb = load_workbook(uploaded_file)
    sheet_names = wb.sheetnames
    attribute_sheet = sheet_names[0]
    attributes = pd.read_excel(uploaded_file, sheet_name=attribute_sheet)
    attrbutes = attributes.astype(str)
    attrbutes.colunms = ['NodeID'] + attributes.columns[1:]

    edge_sheet = st.sidebar.selectbox("Select Network Sheet", sheet_names[1:])
    edges = pd.read_excel(uploaded_file, sheet_name=edge_sheet)
    edges = edges.astype(str)
    edges.columns = ['source', 'target']
    for col in edges.columns:
        edges[col] = edges[col].str.strip()
    edges.dropna(inplace=True)
    edges.drop_duplicates(inplace=True)

    return attributes, edges


# Define a function to generate the report
def generate_report(html_content, filename):
    # Create a PDF file
    pdf_file = HTML(string=html_content).write_pdf()

    # Save PDF file
    with open(filename, "wb") as file:
     st.write(pdf_file, format="pdf")
     # Provide download link
    with open(filename, "rb") as file:
     st.download_button(label="Download Report", data=file, file_name=filename, mime="application/pdf")


if __name__ == "__main__":

    st.title("Network Analysis App")
    st.sidebar.title("Options")
    

    # File Upload Functionality
    st.sidebar.title("Upload File")
    uploaded_file = st.sidebar.file_uploader("Upload a CSV or an Excel File", type=["csv", "xlsx"])

    

    

    # Download report
    html_content = """
    <h1>My Streamlit App</h1>
    <p>This is the content of my Streamlit app.</p>
    """
    report_button_id = "report_button"
    if st.sidebar.button("Generate Report"):
        generate_report(html_content, "report.pdf")

    
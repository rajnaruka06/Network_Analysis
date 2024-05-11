import streamlit as st
import pandas as pd
import networkx as nx
from pyvis.network import Network
from networkx.algorithms import community
import matplotlib.colors as mcolors
import numpy as np
from openpyxl import load_workbook
import tempfile
import subprocess

from PIL import Image
import pyautogui

import os
os.add_dll_directory(r"C:\Program Files\GTK3-Runtime Win64\bin")
from datetime import datetime
import tempfile



@st.cache_data()
def install_r_packages():
    r_script_path = 'install_ergm.R'
    r_command = ['Rscript', r_script_path]
    process = subprocess.Popen(r_command, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
    output, error = process.communicate()
    return output, error

def create_graph(network_df):
    graph = nx.Graph()
    for i in range(len(network_df)):
        graph.add_edge(network_df['source'][i], network_df['target'][i])
    
    return graph

def calculate_network_statistics(graph):
    
    connected_flag = nx.is_connected(graph)
    if not connected_flag:
        st.warning("The network is not connected. Some network Statistics can not be claculated.")

    number_of_nodes = graph.number_of_nodes()
    number_of_edges = graph.number_of_edges()
    average_degree = sum(dict(graph.degree()).values()) / graph.number_of_nodes()
    density = nx.density(graph)
    clustering_coefficient = nx.average_clustering(graph)
    average_shortest_path_length = nx.average_shortest_path_length(graph) if connected_flag else "Can't calculate Graph is not connected"
    diameter = nx.diameter(graph) if connected_flag else "Can't calculate Graph is not connected"
    degree_centrality = nx.degree_centrality(graph)
    closeness_centrality = nx.closeness_centrality(graph)
    betweenness_centrality = nx.betweenness_centrality(graph)
    eigenvector_centrality = nx.eigenvector_centrality(graph, max_iter=500)
    pagerank = nx.pagerank(graph, max_iter=500)
    hits = nx.hits(graph)

    communities = community.louvain_communities(graph)
    number_of_communities = len(communities)
    community_sizes = [len(community) for community in communities]
    largest_community = max(community_sizes)
    smallest_community = min(community_sizes)
    modularity = community.modularity(graph, communities)

    statstics = {
        "Number of Nodes": number_of_nodes,
        "Number of Edges": number_of_edges,
        "Average Degree": average_degree,
        "Density": density,
        "Clustering Coefficient": clustering_coefficient,
        "Average Shortest Path Length": average_shortest_path_length,
        "Diameter": diameter,
        "Degree Centrality": degree_centrality,
        "Closeness Centrality": closeness_centrality,
        "Betweenness Centrality": betweenness_centrality,
        "Eigenvector Centrality": eigenvector_centrality,
        "PageRank": pagerank,
        "HITS Hub Scores": hits[0],
        "HITS Authority Scores": hits[1],
        'Number of communities': number_of_communities,
        'Community with the largest size': largest_community,
        'Community with the smallest size': smallest_community,
        'Modularity': modularity,
        'Communities': communities
    }
    return statstics

def create_network_visualization(graph, selected_visual_metrics, network_statistics, annotate):
    nt = Network()
    nt.from_nx(graph)
    pos = nx.spectral_layout(graph)

    max_metric_value = max(network_statistics[selected_visual_metrics].values())
    min_metric_value = min(network_statistics[selected_visual_metrics].values())

    for node, coords in pos.items():
        for node_dict in nt.nodes:
            if node_dict['id'] == node:
                node_dict['x'] = coords[0]
                node_dict['y'] = coords[1]
                metric_value = network_statistics[selected_visual_metrics][node]
                if annotate:
                    node_label = f"ID: {node}"
                    node_label += f", {selected_visual_metrics}: {metric_value:.2f}"
                    node_dict['label'] = node_label
                color_density = (metric_value - min_metric_value) / (max_metric_value - min_metric_value)
                node_dict['color'] = f"rgba(0, 0, 255, {color_density})"
                break
    
    output_dir = "./app_graph_layout.html"
    nt.save_graph(output_dir)
    
    return output_dir

def create_community_visualization(graph, network_statistics):
    nt = Network()
    nt.from_nx(graph)
    pos = nx.spectral_layout(graph)

    ## Community coloring
    number_of_communities = network_statistics['Number of communities']
    communities = network_statistics['Communities']
    pastel_cmap = mcolors.LinearSegmentedColormap.from_list(
        'pastel_cmap', 
        np.random.rand(number_of_communities, 4), 
        N=number_of_communities
    )
    pastel_colors = pastel_cmap(np.linspace(0, 1, number_of_communities))
    node_comunity_color_map = dict()
    for idx, community in enumerate(communities):
        color = mcolors.to_hex(pastel_colors[idx])
        for node in community:
            node_comunity_color_map[node] = color
    
    for node, coords in pos.items():
        for node_dict in nt.nodes:
            if node_dict['id'] == node:
                node_dict['x'] = coords[0]
                node_dict['y'] = coords[1]
                node_dict['color'] = node_comunity_color_map[node]
                break
    
    output_dir = "./app_community_layout.html"
    nt.save_graph(output_dir)
    return output_dir


def _read_csv(upload_file):
    df = pd.read_csv(upload_file)
    df.columns = ['source', 'target']
    df = df.astype(str)
    for col in df.columns:
        df[col] = df[col].str.strip()
    df.dropna(inplace=True)
    df.drop_duplicates(inplace=True)
    df = df[~(df['source'] == df['target'])]
    df.reset_index(drop=True, inplace=True)

    return df

def _read_excel(uploaded_file):
    wb = load_workbook(uploaded_file)
    sheet_names = wb.sheetnames
    attribute_sheet = sheet_names[0]
    attributes = pd.read_excel(uploaded_file, sheet_name=attribute_sheet)
    attributes = attributes.astype(str)
    attributes.columns = ['NodeID'] +  list(attributes.columns[1:])

    edge_sheet = st.sidebar.selectbox("Select Network", sheet_names[1:])
    edges = pd.read_excel(uploaded_file, sheet_name=edge_sheet)
    edges = edges.astype(str)
    edges.columns = ['source', 'target']
    for col in edges.columns:
        edges[col] = edges[col].str.strip()
    edges.dropna(inplace=True)
    edges.drop_duplicates(inplace=True)
    edges  = edges[~(edges['source'] == edges['target'])]
    edges.reset_index(drop=True, inplace=True)

    return attributes, edges

def _show_ergm_report(summary_text, edges_only=False):
    call_section, results_section = summary_text.split("\n\nMaximum Likelihood Results:")
    results_lines = results_section.splitlines()
    headers = ['Estimate', 'Std. Error', 'MCMC %', 'z value', 'Pr(>|z|)']
    data = []
    for line in results_lines[3:-7]:
        row = []
        for val in line.split()[1:-1]:
            try:
                row.append(float(val))
            except ValueError:
                try:
                    row.append(float(val[1:]))
                except ValueError:
                    pass
            else:
                continue
        data.append(row)
    
    dynamic_indexes = ['edges', f'nodematch-{selected_attribute}'] if not edges_only else ['edges']
    summary_df = pd.DataFrame(data, columns=headers, index=dynamic_indexes)

    null_deviance_line = results_lines[-4].split(": ")[1].split()
    null_deviance_value, null_deviance_df  = null_deviance_line[0], null_deviance_line[2]

    residual_deviance_line = results_lines[-3].split(": ")[1].split()
    residual_deviance_value, residual_deviance_df = residual_deviance_line[0], residual_deviance_line[2]

    aic_bic_line = results_lines[-1].split(": ")
    aic_value = aic_bic_line[1].split()[0]
    bic_value = aic_bic_line[2].split()[0]
    
    st.table(summary_df)
    
    st.write("Null Deviance:", null_deviance_value, f"on {null_deviance_df} degrees of freedom")
    st.write("Residual Deviance:", residual_deviance_value, f"on {residual_deviance_df} degrees of freedom")
    st.write("AIC:", aic_value)
    st.write("BIC:", bic_value)

def save_screenshot(region=None):
    download_folder = os.path.expanduser("~/Downloads")
    screenshot_name = f"screenshot_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.png"
    screenshot_path = os.path.join(download_folder, screenshot_name)
    
    if region:
        pyautogui.screenshot(screenshot_path, region=region)
    else:
        pyautogui.screenshot(screenshot_path)
    
    return screenshot_path




if __name__ == "__main__":

    st.set_page_config(page_title="Network Analysis App", page_icon="📊", layout="wide")
    output, error = install_r_packages()

    # st.write(os.listdir("./r_packages"))

    st.title("Network Analysis App")
    st.sidebar.title("Options")

    # File Upload FUnctionality
    st.sidebar.title("Upload File")
    uploaded_file = st.sidebar.file_uploader("Upload a CSV or an Excel File", type=["csv", "xlsx"])

    if uploaded_file is not None:
        with st.spinner("Uploading file..."):
            if uploaded_file.name.endswith(".csv"):
                network_df = _read_csv(uploaded_file)
                attribute_df = pd.DataFrame(columns=['NodeID'])
            else:
                attribute_df, network_df =  _read_excel(uploaded_file)
        st.sidebar.success("File Uploaded Successfully!")

        graph = create_graph(network_df)
        network_statistics = calculate_network_statistics(graph)

        # Network Visualization and Metrics
        st.sidebar.title("Select Visual Metrics")
        visual_metrics_list = ("Degree Centrality", "Closeness Centrality", "Betweenness Centrality", "Eigenvector Centrality",
                            "PageRank", "HITS Hub Scores", "HITS Authority Scores")
        selected_visual_metrics = st.sidebar.selectbox("Select Visual Metrics", visual_metrics_list)

        st.header("Network Visualization")
        st.markdown("Zoom in/out, Drag or select to see individual node and its attributes.")
        annotate = st.checkbox("Annotate Nodes with Visual Metric Value")
        viz_path = create_network_visualization(graph, selected_visual_metrics, network_statistics, annotate)
        

        # Community Visualization
        show_community_visualization = st.checkbox("Show Community Visualization")
        if show_community_visualization:
            st.header("Community Visualization")
            st.markdown("Zoom in/out, Drag or select to see individual node and its attributes.")
            community_viz_path = create_community_visualization(graph, network_statistics)
            st.components.v1.html(open(community_viz_path, 'r', encoding='utf-8').read(), height=800)

        # Network Statistics
        st.sidebar.title("Select Network Statistics")
        metrics_list = ("Number of Nodes", "Number of Edges", "Average Degree", "Density", "Clustering Coefficient", "Average Shortest Path Length", "Diameter"
                        , 'Number of communities', 'Community with the largest size', 'Community with the smallest size', 'Modularity')
        selected_metrics = st.sidebar.multiselect("Select Metrics", metrics_list)
        
        if selected_metrics: 
            st.header("Network Analysis Metrics")
        for metric in selected_metrics:
            value = network_statistics.get(metric)
            if isinstance(value, (int, float)):
                st.write(f"**{metric}:** {value:.2f}")
            else:
                st.write(f"**{metric}:** {value}")
        
        # Statistical Modeling
        st.sidebar.title("Select Statistical Model")
        selected_model = st.sidebar.radio("Choose Model", ("ERGM", "ALAAM"))
        selected_attribute =  st.sidebar.selectbox("Select Attribute", attribute_df.columns[1:])
        


        # Download report
        st.sidebar.title("Download Report")
        report_button = st.sidebar.button("Download Report")
        full_screen = st.sidebar.checkbox("Capture Entire Screen", value=True, key="full_screen_checkbox")
        
    
    if report_button:
     screenshot_path = None  # Initialize screenshot_path variable

     if full_screen:
        screenshot_path = save_screenshot()
     else:
        # Wait for the user to press Enter to capture the selected region
        st.text("Click and drag to select the area, then press Enter to capture.")
        st.text("Once selected, press Enter to capture the screenshot.")
        st.text("Press 'q' to quit.")
        
        if st.button("Capture Screenshot"):
            st.write("Capturing screenshot...")
            # Get the position and size of the selected region
            x, y, width, height = pyautogui.locateAndSizeOnScreen(None, confidence=0.8)
            # Capture the screenshot of the selected region
            screenshot_path = save_screenshot(region=(x, y, width, height))
    
     if screenshot_path is not None:
        if os.path.exists(screenshot_path):
            st.success("Screenshot saved successfully to download folder.")
        else:
            st.error("Failed to save screenshot.")
else:
    st.warning("Please Upload a File")
import streamlit as st
import pandas as pd
import networkx as nx
from pyvis.network import Network
from networkx.algorithms import community
import matplotlib.colors as mcolors
import numpy as np
from openpyxl import load_workbook
import subprocess


import rpy2.robjects as ro
from rpy2.robjects import pandas2ri
from rpy2.robjects.conversion import localconverter

@st.cache_data()
def install_r_packages():
    r_script_path = 'install_ergm.R'
    r_command = ['Rscript', r_script_path]
    process = subprocess.Popen(r_command, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
    output, error = process.communicate()
    return output, error

def create_graph(network_df):
    graph = nx.Graph()
    for i in range(len(network_df)):
        graph.add_edge(network_df['source'][i], network_df['target'][i])
    
    return graph

def calculate_network_statistics(graph):
    
    connected_flag = nx.is_connected(graph)
    if not connected_flag:
        st.warning("The network is not connected. Some network Statistics can not be claculated.")

    number_of_nodes = graph.number_of_nodes()
    number_of_edges = graph.number_of_edges()
    average_degree = sum(dict(graph.degree()).values()) / graph.number_of_nodes()
    density = nx.density(graph)
    clustering_coefficient = nx.average_clustering(graph)
    average_shortest_path_length = nx.average_shortest_path_length(graph) if connected_flag else "Can't calculate Graph is not connected"
    diameter = nx.diameter(graph) if connected_flag else "Can't calculate Graph is not connected"
    degree_centrality = nx.degree_centrality(graph)
    closeness_centrality = nx.closeness_centrality(graph)
    betweenness_centrality = nx.betweenness_centrality(graph)
    eigenvector_centrality = nx.eigenvector_centrality(graph, max_iter=500)
    pagerank = nx.pagerank(graph, max_iter=500)
    hits = nx.hits(graph)

    communities = community.louvain_communities(graph)
    number_of_communities = len(communities)
    community_sizes = [len(community) for community in communities]
    largest_community = max(community_sizes)
    smallest_community = min(community_sizes)
    modularity = community.modularity(graph, communities)

    statstics = {
        "Number of Nodes": number_of_nodes,
        "Number of Edges": number_of_edges,
        "Average Degree": average_degree,
        "Density": density,
        "Clustering Coefficient": clustering_coefficient,
        "Average Shortest Path Length": average_shortest_path_length,
        "Diameter": diameter,
        "Degree Centrality": degree_centrality,
        "Closeness Centrality": closeness_centrality,
        "Betweenness Centrality": betweenness_centrality,
        "Eigenvector Centrality": eigenvector_centrality,
        "PageRank": pagerank,
        "HITS Hub Scores": hits[0],
        "HITS Authority Scores": hits[1],
        'Number of communities': number_of_communities,
        'Community with the largest size': largest_community,
        'Community with the smallest size': smallest_community,
        'Modularity': modularity,
        'Communities': communities
    }
    return statstics

def create_network_visualization(graph, selected_visual_metrics, network_statistics, annotate):
    nt = Network()
    nt.from_nx(graph)
    pos = nx.spectral_layout(graph)

    max_metric_value = max(network_statistics[selected_visual_metrics].values())
    min_metric_value = min(network_statistics[selected_visual_metrics].values())

    for node, coords in pos.items():
        for node_dict in nt.nodes:
            if node_dict['id'] == node:
                node_dict['x'] = coords[0]
                node_dict['y'] = coords[1]
                metric_value = network_statistics[selected_visual_metrics][node]
                if annotate:
                    node_label = f"ID: {node}"
                    node_label += f", {selected_visual_metrics}: {metric_value:.2f}"
                    node_dict['label'] = node_label
                color_density = (metric_value - min_metric_value) / (max_metric_value - min_metric_value)
                node_dict['color'] = f"rgba(0, 0, 255, {color_density})"
                break
    
    output_dir = "./app_graph_layout.html"
    nt.save_graph(output_dir)
    return output_dir

def create_community_visualization(graph, network_statistics):
    nt = Network()
    nt.from_nx(graph)
    pos = nx.spectral_layout(graph)

    ## Community coloring
    number_of_communities = network_statistics['Number of communities']
    communities = network_statistics['Communities']
    pastel_cmap = mcolors.LinearSegmentedColormap.from_list(
        'pastel_cmap', 
        np.random.rand(number_of_communities, 4), 
        N=number_of_communities
    )
    pastel_colors = pastel_cmap(np.linspace(0, 1, number_of_communities))
    node_comunity_color_map = dict()
    node_community_map = {}
    for idx, community in enumerate(communities):
        color = mcolors.to_hex(pastel_colors[idx])
        for node in community:
            node_comunity_color_map[node] = color
            node_community_map[node] = idx
    
    annotate_nodes = st.checkbox("Annotate Nodes with Community Number")

    for node, coords in pos.items():
        for node_dict in nt.nodes:
            if node_dict['id'] == node:
                node_dict['x'] = coords[0]
                node_dict['y'] = coords[1]
                node_dict['color'] = node_comunity_color_map[node]
                if annotate_nodes:
                    node_label = f"ID: {node}"
                    node_label += f", Comm: {node_community_map[node]}"
                    node_dict['label'] = node_label
                break
    
    output_dir = "./app_community_layout.html"
    nt.save_graph(output_dir)
    return output_dir

def perform_analysis(network_df, attribute_df, selected_attribute, edges_only, output_file_path, gof_output_file_path, model_type):
    if model_type == 'bernoulli':
        summary_text, gof_summary_text = perform_ergm_analysis(network_df, attribute_df, selected_attribute, edges_only, output_file_path, gof_output_file_path)
    elif model_type == 'node_match':
        if edges_only:
            st.error("ERGM Node Match is not Supported for edges only network")
            return None, None
        summary_text, gof_summary_text = perform_ergm_analysis(network_df, attribute_df, selected_attribute, edges_only, output_file_path, gof_output_file_path)
    elif model_type == 'node_covariate':
        if edges_only:
            st.error("ERGM Node Covariate Analysis is not supported for edges only network")
            return None, None
        summary_text, gof_summary_text = perform_alaam_analysis(network_df, attribute_df, selected_attribute, edges_only, output_file_path, gof_output_file_path)
    else:
        st.error("Invalid Model Type")
        return None, None
    return summary_text, gof_summary_text

def perform_ergm_analysis(network_df, attribute_df, selected_attribute, edges_only=False, output_file_path="ergm_analysis_results.txt", gof_output_file_path = "ergm_gof_results.txt"):
    
    ## Bernoulli Model for edges only
    if edges_only:

        pandas2ri.activate()
        with localconverter(ro.default_converter + pandas2ri.converter):
            r_net_data = ro.conversion.py2rpy(network_df)

        ro.globalenv['df'] = r_net_data
        
        ro.r(f'''
            library(network, lib.loc="./r_packages")
            library(ergm, lib.loc="./r_packages")
            df$Source <- as.character(df$source)
            df$Target <- as.character(df$target)
            net <- network::network(df, directed = TRUE, loops = FALSE)

            formula <- "net ~ edges"
            ergm_model <- ergm::ergm(as.formula(formula))
            summary_ergm <- summary(ergm_model)
            writeLines(capture.output(summary_ergm), "{output_file_path}")

            gof_results <- gof(ergm_model, GOF=~odegree+idegree)
            writeLines(capture.output(gof_results), "{gof_output_file_path}")

            ''')
    
    ## Node Match Model for edges and attribute
    else:
        
        attribute_df = attribute_df[['NodeID', selected_attribute]]
        attribute_df.dropna(subset=[selected_attribute], inplace=True)

        net_data = pd.merge(network_df, attribute_df, left_on='source', right_on='NodeID', how='left')
        net_data.drop(columns=['NodeID'], inplace=True)
        
        pandas2ri.activate()
        with localconverter(ro.default_converter + pandas2ri.converter):
            r_net_data = ro.conversion.py2rpy(net_data)

        ro.globalenv['df'] = r_net_data
        ro.globalenv['selected_attribute'] = selected_attribute
        ro.r(f'''
            library(network, lib.loc="./r_packages")
            library(ergm, lib.loc="./r_packages")

            net <- network::network(df, vertex.attr = list({selected_attribute} = df${selected_attribute}), directed = TRUE, loops = FALSE)
            formula <- paste("net ~ edges + nodematch('", "{selected_attribute}", "', diff = FALSE)", sep="")

            ergm_model <- ergm::ergm(as.formula(formula))
            summary_ergm <- summary(ergm_model)
                
            writeLines(capture.output(summary_ergm), "{output_file_path}")

            gof_results <- gof(ergm_model, GOF=~odegree+idegree)
            writeLines(capture.output(gof_results), "{gof_output_file_path}")
            ''')

        
    with open(output_file_path, 'r') as f:
        summary_text = f.read().strip()
    # st.download_button(label="Download ERGM Analysis Results", data=summary_text, mime="text/plain", file_name="ergm_analysis_results.txt")

    with open(gof_output_file_path, 'r') as f:
        gof_summary_text = f.read().strip()

    return summary_text, gof_summary_text

def perform_alaam_analysis(network_df, attribute_df, selected_attribute, edges_only=False, output_file_path="alaam_analysis_results.txt", gof_output_file_path = "alaam_gof_results.txt"):

    if edges_only:
        st.error("ERGM Node Covariate Analysis is not supported for edges only network")
    
    ## Node Covariate Model for edges and attribute
    else:
        if attribute_df[selected_attribute].isna().any():
            st.error("ERGM Node Covariate Analysis is not supported for missing values in the selected attribute")
            return None
        try:
            attribute_df[selected_attribute] = attribute_df[selected_attribute].astype(int)
        except:
            st.error("ERGM Node Covariate Analysis is only supported for numeric attributes")
            return None
        
        attribute_df = attribute_df[['NodeID', selected_attribute]]
        attribute_df.dropna(subset=[selected_attribute], inplace=True)

        net_data = pd.merge(network_df, attribute_df, left_on='source', right_on='NodeID', how='left')
        net_data.drop(columns=['NodeID'], inplace=True)
        
        pandas2ri.activate()
        with localconverter(ro.default_converter + pandas2ri.converter):
            r_net_data = ro.conversion.py2rpy(net_data)

        ro.globalenv['df'] = r_net_data
        ro.globalenv['selected_attribute'] = selected_attribute
        ro.r(f'''
            library(network, lib.loc="./r_packages")
            library(ergm, lib.loc="./r_packages")

            net <- network::network(df, vertex.attr = list({selected_attribute} = df${selected_attribute}), directed = TRUE, loops = FALSE)
            formula <- paste("net ~ edges + nodecov('", "{selected_attribute}", "')", sep="")
            
            alaam_model <- ergm::ergm(as.formula(formula))
            summary_alaam <- summary(alaam_model)
                
            writeLines(capture.output(summary_alaam), "{output_file_path}")

            gof_results <- gof(ergm_model, GOF=~odegree+idegree)
            writeLines(capture.output(gof_results), "{gof_output_file_path}")
            
            ''')

        
    with open(output_file_path, 'r') as f:
        summary_text = f.read().strip()
    # st.download_button(label="Download ALAAM Analysis Results", data=summary_text, mime="text/plain", file_name="alaam_analysis_results.txt")

    with open(gof_output_file_path, 'r') as f:
        gof_summary_text = f.read().strip()

    return summary_text, gof_summary_text

def _read_csv(upload_file):
    df = pd.read_csv(upload_file)
    df.columns = ['source', 'target']
    df = df.astype(str)
    for col in df.columns:
        df[col] = df[col].str.strip()
    df.dropna(inplace=True)
    df.drop_duplicates(inplace=True)
    df = df[~(df['source'] == df['target'])]
    df.reset_index(drop=True, inplace=True)

    return df

def _read_excel(uploaded_file):
    wb = load_workbook(uploaded_file)
    sheet_names = wb.sheetnames
    attribute_sheet = sheet_names[0]
    attributes = pd.read_excel(uploaded_file, sheet_name=attribute_sheet)
    attributes = attributes.astype(str)
    attributes.columns = ['NodeID'] +  list(attributes.columns[1:])

    edge_sheet = st.sidebar.selectbox("Select Network", sheet_names[1:])
    edges = pd.read_excel(uploaded_file, sheet_name=edge_sheet)
    edges = edges.astype(str)
    edges.columns = ['source', 'target']
    for col in edges.columns:
        edges[col] = edges[col].str.strip()
    edges.dropna(inplace=True)
    edges.drop_duplicates(inplace=True)
    edges  = edges[~(edges['source'] == edges['target'])]
    edges.reset_index(drop=True, inplace=True)

    return attributes, edges

def _show_ergm_report(summary_text, edges_only=False):
    call_section, results_section = summary_text.split("\n\nMaximum Likelihood Results:")
    results_lines = results_section.splitlines()
    headers = ['Estimate', 'Std. Error', 'MCMC %', 'z value', 'Pr(>|z|)']
    data = []
    for line in results_lines[3:-7]:
        row = []
        for val in line.split()[1:-1]:
            try:
                row.append(float(val))
            except ValueError:
                try:
                    row.append(float(val[1:]))
                except ValueError:
                    pass
            else:
                continue
        data.append(row)
    
    dynamic_indexes = ['edges', f'nodematch-{selected_attribute}'] if not edges_only else ['edges']
    summary_df = pd.DataFrame(data, columns=headers, index=dynamic_indexes)

    null_deviance_line = results_lines[-4].split(": ")[1].split()
    null_deviance_value, null_deviance_df  = null_deviance_line[0], null_deviance_line[2]

    residual_deviance_line = results_lines[-3].split(": ")[1].split()
    residual_deviance_value, residual_deviance_df = residual_deviance_line[0], residual_deviance_line[2]

    aic_bic_line = results_lines[-1].split(": ")
    aic_value = aic_bic_line[1].split()[0]
    bic_value = aic_bic_line[2].split()[0]
    
    st.table(summary_df)
    
    st.write("Null Deviance:", null_deviance_value, f"on {null_deviance_df} degrees of freedom")
    st.write("Residual Deviance:", residual_deviance_value, f"on {residual_deviance_df} degrees of freedom")
    st.write("AIC:", aic_value)
    st.write("BIC:", bic_value)

def _show_alaam_report(summary_text, edges_only=False):
    call_section, results_section = summary_text.split("\n\nMaximum Likelihood Results:")
    results_lines = results_section.splitlines()
    headers = ['Estimate', 'Std. Error', 'MCMC %', 'z value', 'Pr(>|z|)']
    data = []
    for line in results_lines[3:-7]:
        row = []
        for val in line.split()[1:-1]:
            try:
                row.append(float(val))
            except ValueError:
                try:
                    row.append(float(val[1:]))
                except ValueError:
                    pass
            else:
                continue
        data.append(row)
    
    dynamic_indexes = ['edges', f'nodecov-{selected_attribute}'] if not edges_only else ['edges']
    summary_df = pd.DataFrame(data, columns=headers, index=dynamic_indexes)

    null_deviance_line = results_lines[-4].split(": ")[1].split()
    null_deviance_value, null_deviance_df  = null_deviance_line[0], null_deviance_line[2]

    residual_deviance_line = results_lines[-3].split(": ")[1].split()
    residual_deviance_value, residual_deviance_df = residual_deviance_line[0], residual_deviance_line[2]

    aic_bic_line = results_lines[-1].split(": ")
    aic_value = aic_bic_line[1].split()[0]
    bic_value = aic_bic_line[2].split()[0]
    
    st.table(summary_df)
    
    st.write("Null Deviance:", null_deviance_value, f"on {null_deviance_df} degrees of freedom")
    st.write("Residual Deviance:", residual_deviance_value, f"on {residual_deviance_df} degrees of freedom")
    st.write("AIC:", aic_value)
    st.write("BIC:", bic_value)

def _show_gof_report(gof_summary_text, edges_only=False):
    st.header("Goodness of Fit Results")
    # st.download_button(label="Download GOF Results", data=gof_summary_text, mime="text/plain", file_name="gof_results.txt", key="gof_results")

    lines = gof_summary_text.split("\n")
    headers = lines[2].split()
    headers = ['degree'] + headers[:4] + headers[5:]
    for idx, line in enumerate(lines):
        if line.startswith("Goodness-of-fit for in-degree"):
            odegree_end = idx - 1
        elif line.startswith("Goodness-of-fit for model statistics"):
            indegree_end = idx - 1
    out_degree_dof_data = lines[3:odegree_end]
    in_degree_dof_data = lines[odegree_end + 4:indegree_end]
    network_data = lines[-2:] if not edges_only else lines[-1:]
    
    out_degree_dof_data = [line.split() for line in out_degree_dof_data]
    in_degree_dof_data = [line.split() for line in in_degree_dof_data]
    network_data = [line.split() for line in network_data]
    if edges_only: 
        network_data[0] = ['edges'] + network_data[0]
    
    out_degree_df = pd.DataFrame(out_degree_dof_data, columns=headers)
    in_degree_df = pd.DataFrame(in_degree_dof_data, columns=headers)
    network_df = pd.DataFrame(network_data, columns=['model statistic'] + headers[1:])

    st.write("Out Degree Goodness of Fit")
    st.table(out_degree_df)
    st.write("In Degree Goodness of Fit")
    st.table(in_degree_df)
    st.write("Network Goodness of Fit")
    st.table(network_df)

    return out_degree_df, in_degree_df, network_df

if __name__ == "__main__":

    st.set_page_config(page_title="Network Analysis App", page_icon="📊", layout="wide")
    output, error = install_r_packages()

    st.title("Network Analysis App")
    st.sidebar.title("Options")

    ## File Upload FUnctionality
    st.sidebar.title("Upload File")
    supported_formats = ["csv", "xlsx"]
    uploaded_file = st.sidebar.file_uploader("Upload a CSV or an Excel File", type=supported_formats)

    if uploaded_file is not None:
        with st.spinner("Uploading file..."):
            if uploaded_file.name.endswith(".csv"):
                network_df = _read_csv(uploaded_file)
                attribute_df = pd.DataFrame(columns=['NodeID'])
            else:
                attribute_df, network_df =  _read_excel(uploaded_file)
        st.sidebar.success("File Uploaded Successfully!")

        graph = create_graph(network_df)
        network_statistics = calculate_network_statistics(graph)

        ## Network Visualization and Metrics
        st.sidebar.title("Select Visual Metrics")
        visual_metrics_list = ("Degree Centrality", "Closeness Centrality", "Betweenness Centrality", "Eigenvector Centrality",
                            "PageRank", "HITS Hub Scores", "HITS Authority Scores")
        selected_visual_metrics = st.sidebar.selectbox("Select Visual Metrics", visual_metrics_list)

        st.header("Network Visualization")
        st.markdown("Zoom in/out, Drag or select to see individual node and its attributes.")
        annotate = st.checkbox("Annotate Nodes with Visual Metric Value")
        viz_path = create_network_visualization(graph, selected_visual_metrics, network_statistics, annotate)
        with open(viz_path, 'r', encoding='utf-8') as f:
            html_content = open(viz_path, 'r', encoding='utf-8').read()
            st.download_button(label="Download Network Visualization", data=html_content, mime="text/html", file_name="network_visualization.html", key="network_viz")
            st.components.v1.html(html_content, height=600)

        ## Community Visualization
        show_community_visualization = st.checkbox("Show Community Visualization")
        if show_community_visualization:
            st.header("Community Visualization")
            st.markdown("Zoom in/out, Drag or select to see individual node and its attributes.")
            community_viz_path = create_community_visualization(graph, network_statistics)
            html_content = open(community_viz_path, 'r', encoding='utf-8').read()
            st.download_button(label="Download Community Visualization", data=html_content, mime="text/html", file_name="community_visualization.html", key="community_viz")
            community_metrics = ("Number of communities", "Community with the largest size", "Community with the smallest size", "Modularity")
            st.components.v1.html(html_content, height=800)

        ## Network Statistics
        st.sidebar.title("Select Network Statistics")
        metrics_list = ("Number of Nodes", "Number of Edges", "Average Degree", "Density", "Clustering Coefficient", "Average Shortest Path Length", "Diameter"
                        , 'Number of communities', 'Community with the largest size', 'Community with the smallest size', 'Modularity')
        selected_metrics = st.sidebar.multiselect("Select Metrics", metrics_list)
        
        if selected_metrics: 
            st.header("Network Analysis Metrics")
        for metric in selected_metrics:
            value = network_statistics.get(metric)
            if isinstance(value, (int, float)):
                st.write(f"**{metric}:** {value:.2f}")
            else:
                st.write(f"**{metric}:** {value}")
        
        ## Statistical Modeling
        st.sidebar.title("Select Statistical Model")
        # selected_model = st.sidebar.radio("Choose Model", ("ERGM", "ALAAM"))
        selected_model = st.sidebar.selectbox("Choose Model", ("bernoulli", "node_match", "node_covariate"))
        selected_attribute =  st.sidebar.selectbox("Select Attribute", attribute_df.columns[1:])

        st.header(f"{selected_model} ERGM Analysis Summary")
        edges_only=uploaded_file.name.endswith(".csv")
        file_path = "analysis_result.txt"
        gof_file_path = "gof_results.txt"
        with st.spinner(f"Performing {selected_model} ERGM Analysis..."):
            summary_text, gof_summary_text = perform_analysis(network_df, attribute_df, selected_attribute, edges_only=edges_only, output_file_path=file_path, gof_output_file_path=gof_file_path, model_type=selected_model)
        
        ## Manual labour to display ERGM summary
        if summary_text is not None:
            if selected_model in ['bernoulli', 'node_match']:
                _show_ergm_report(summary_text, edges_only = edges_only)
            elif selected_model == 'node_covariate':
                _show_alaam_report(summary_text, edges_only = edges_only)
        else:
            st.error("An error occurred during analysis")
        
        show_gof = False
        if gof_summary_text is not None:
            show_gof = st.checkbox("Show Goodness Of Fit Results")
            if show_gof:
                out_degree_df, in_degree_df, network_dof_df = _show_gof_report(gof_summary_text, edges_only=edges_only)

        # if selected_model == "ERGM":
        #     st.header("ERGM Analysis Summary")
        #     edges_only=uploaded_file.name.endswith(".csv")
        #     ergm_file_path = "ergm_analysis_results.txt"
        #     gof_file_path = "ergm_gof_results.txt"
        #     with st.spinner("Performing ERGM Analysis..."):
        #         summary_text, gof_summary_text = perform_ergm_analysis(network_df, attribute_df,  selected_attribute, edges_only=edges_only, output_file_path=ergm_file_path, gof_output_file_path=gof_file_path)
                        
        #     ## Manual labour to display ERGM summary
        #     if summary_text is not None:
        #         _show_ergm_report(summary_text, edges_only = edges_only)
        #     if gof_summary_text is not None:
        #         show_gof = st.checkbox("Show Goodness Of Fit Results")
        #         if show_gof:
        #             out_degree_df, in_degree_df, network_dof_df = _show_gof_report(gof_summary_text, edges_only=edges_only)
        #     else:
        #         st.error("An error occurred during ERGM analysis")
        
        # elif selected_model == "ALAAM":
        #     st.header("ALAAM Analysis")
        #     st.warning("This is ERGM + Node Covariate Model.")
        #     if uploaded_file.name.endswith(".csv"):
        #         st.warning("ALAAM Analysis is not supported for edges only network")    
        #     else:
        #         alaam_file_path = "alaam_analysis_results.txt"
        #         gof_file_path = "alaam_gof_results.txt"
        #         summary_text, gof_summary_text = perform_alaam_analysis(network_df, attribute_df, selected_attribute, output_file_path=alaam_file_path, gof_output_file_path=gof_file_path)
        #         if summary_text is not None:
        #             _show_alaam_report(summary_text)
        #         if gof_summary_text is not None:
        #             show_gof = st.checkbox("Show Goodness Of Fit Results")
        #             if show_gof:
        #                 out_degree_df, in_degree_df, network_dof_df = _show_gof_report(gof_summary_text)
        #         else:
        #             st.error("An error occurred during ALAAM analysis")


        ## Download report
        report_df = network_df.copy()
        report_df.drop(columns=['target'], inplace=True)
        
        node_community_map = {}
        for community_num, community_list in enumerate(network_statistics['Communities']):
            for node in community_list:
                node_community_map[node] = community_num

        report_df['Community'] = report_df['source'].map(node_community_map)
        node_lev_statistics = ["Degree Centrality", "Closeness Centrality", "Betweenness Centrality", "Eigenvector Centrality", "PageRank", "HITS Hub Scores", "HITS Authority Scores"]
        for stat in node_lev_statistics:
            report_df[stat] = report_df['source'].map(network_statistics[stat])
        
        report_df.columns = ['Node'] + report_df.columns[1:].tolist()

        writer = pd.ExcelWriter('Network_Analysis.xlsx', engine='xlsxwriter')
        report_df.to_excel(writer, sheet_name='Node_Level_Stats', index=False)
        network_statistics_df = pd.DataFrame.from_dict({stat: [network_statistics[stat]] for stat in metrics_list}, orient='index')
        network_statistics_df.reset_index(inplace=True)
        network_statistics_df.columns = ['Metric', 'Value']
        if network_statistics_df is not None:
            network_statistics_df.to_excel(writer, sheet_name='Network Statistics', index=False)
        
        if summary_text is not None:
            summary_df = pd.DataFrame.from_dict({'Summary': summary_text[summary_text.find('Maximum'):]}, orient='index')
            summary_df.reset_index(inplace=True)
            if summary_df is not None:
                summary_df.to_excel(writer, sheet_name=f'{selected_model} Summary', index=False)

        if show_gof:
            gof_dfs = {'Out Degree': out_degree_df, 'In Degree': in_degree_df, 'Network': network_dof_df}
            for key, df in gof_dfs.items():
                if df is not None:
                    df.to_excel(writer, sheet_name=f'{key} GOF', index=False)

        writer._save()
        with open('Network_Analysis.xlsx', 'rb') as f:
            file_content = f.read()
            ## Give horizontal line
            st.markdown("---")
            st.download_button(label="Download Analysis Report", data=file_content, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", file_name="Network_Analysis.xlsx", key="analysis_report")
            
    else:
        st.warning("Please Upload a File")
        st.header("Instructions for file format")

        st.write(f"**Note:** The app only supports {', '.join(supported_formats)} files.")

        st.subheader("CSV File Format:")
        csv_description = """
        The CSV file should represent a directed network (graph).

        - It must have two columns:
            - Source Node: The starting node for each connection.
            - Target Node: The ending node for each connection.
        """
        st.write(csv_description)

        st.subheader("Excel File Format:")
        excel_description = """
        The Excel file should contain two sheets:

        1. Attributes Sheet:
            - The first column must be the NodeID - should contain unique identifiers for each node.
            - Subsequent columns can contain various attributes for each node.

        2. Network Sheet (Same format as CSV):
            - It must have two columns:
                - Source Node: The starting node for each connection.
                - Target Node: The ending node for each connection.
        """
        st.write(excel_description)